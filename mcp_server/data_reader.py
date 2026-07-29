# -*- coding: utf-8 -*-
"""读取 MediaCrawler 爬取产生的数据文件。"""

from __future__ import annotations

import csv
import json
import os
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional

# MediaCrawler 项目根目录
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 数据存储根目录
DATA_DIR = os.path.join(PROJECT_ROOT, "data")

# MCP 使用短平台代号，部分文件存储实现使用完整英文目录名。
PLATFORM_DATA_DIRS = {
    "dy": "douyin",
    "ks": "kuaishou",
    "wb": "weibo",
}
EXCEL_PLATFORM_DATA_DIRS = {
    **PLATFORM_DATA_DIRS,
    "bili": "bilibili",
}

READABLE_FILE_TYPES = frozenset({"jsonl", "json", "csv", "excel"})
ITEM_TYPES = ("contents", "comments", "creators", "user_actions")
MAX_JSON_FILE_BYTES = 64 * 1024 * 1024


def _get_date_str() -> str:
    """获取当前日期字符串 (YYYY-MM-DD)。"""
    return datetime.now().strftime("%Y-%m-%d")


def _find_data_file_groups(
    platform: str,
    crawler_type: str,
    file_type: str = "jsonl",
    date_str: Optional[str] = None,
    data_root: Optional[str] = None,
) -> Dict[str, List[str]]:
    """返回每类产物的全部候选文件。

    ``data_root`` 用于 MCP 单次运行的隔离目录。指定后不按日期过滤，
    因为该目录本身已经唯一标识一次运行。
    """
    if file_type not in READABLE_FILE_TYPES:
        return {}

    root = Path(data_root or DATA_DIR)
    platform_dir = (
        EXCEL_PLATFORM_DATA_DIRS.get(platform, platform)
        if file_type == "excel"
        else PLATFORM_DATA_DIRS.get(platform, platform)
    )

    if file_type == "excel":
        base_path = root / platform_dir
        if not base_path.exists():
            return {}
        prefix = f"{platform_dir}_{crawler_type}_"
        candidates = sorted(
            (
                str(path)
                for path in base_path.iterdir()
                if path.is_file()
                and path.name.startswith(prefix)
                and path.suffix.lower() == ".xlsx"
            ),
            key=os.path.getmtime,
        )
        return {"workbook": candidates} if candidates else {}

    base_path = root / platform_dir / file_type
    if not base_path.exists():
        return {}

    effective_date = date_str
    if data_root is None and effective_date is None:
        effective_date = _get_date_str()

    files: Dict[str, List[str]] = {}
    for item_type in ITEM_TYPES:
        prefix = f"{crawler_type}_{item_type}_"
        candidates = []
        for path in base_path.iterdir():
            if (
                not path.is_file()
                or not path.name.startswith(prefix)
                or path.suffix.lower() != f".{file_type}"
            ):
                continue
            suffix = path.name[len(prefix):]
            if effective_date and not suffix.startswith(effective_date):
                continue
            candidates.append(str(path))
        if candidates:
            files[item_type] = sorted(candidates, key=os.path.getmtime)
    return files


def find_data_files(
    platform: str,
    crawler_type: str,
    file_type: str = "jsonl",
    date_str: Optional[str] = None,
    data_root: Optional[str] = None,
) -> Dict[str, str]:
    """查找爬取产物，并为兼容旧调用返回每类最新的一个文件。"""
    groups = _find_data_file_groups(
        platform=platform,
        crawler_type=crawler_type,
        file_type=file_type,
        date_str=date_str,
        data_root=data_root,
    )
    return {item_type: paths[-1] for item_type, paths in groups.items() if paths}


def _json_safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _read_excel_file(file_path: str, max_items: int) -> Dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    all_items: List[Dict[str, Any]] = []
    sheets: Dict[str, Dict[str, Any]] = {}
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                continue
            headers = [
                str(value) if value not in (None, "") else f"column_{index + 1}"
                for index, value in enumerate(header_row)
            ]
            sheet_items: List[Dict[str, Any]] = []
            sheet_count = 0
            for row in rows:
                if not any(value is not None for value in row):
                    continue
                sheet_count += 1
                item = {
                    header: _json_safe_value(value)
                    for header, value in zip(headers, row)
                }
                if len(sheet_items) < max_items:
                    sheet_items.append(item)
                if len(all_items) < max_items:
                    all_items.append({"_sheet": worksheet.title, **item})
            sheets[worksheet.title] = {
                "count": sheet_count,
                "returned_count": len(sheet_items),
                "truncated": sheet_count > len(sheet_items),
                "items": sheet_items,
            }
    finally:
        workbook.close()

    total = sum(sheet["count"] for sheet in sheets.values())
    return {
        "count": total,
        "returned_count": len(all_items),
        "truncated": total > len(all_items),
        "items": all_items,
        "sheets": sheets,
    }


def read_data_file(file_path: str, max_items: int = 100) -> Dict[str, Any]:
    """读取单个 JSONL、JSON、CSV 或 Excel 数据文件。"""
    if not os.path.exists(file_path):
        return {
            "count": 0,
            "returned_count": 0,
            "truncated": False,
            "items": [],
        }

    ext = os.path.splitext(file_path)[1].lower()
    items: List[Any] = []
    total = 0
    read_errors: List[str] = []
    invalid_record_count = 0

    try:
        if ext == ".xlsx":
            return _read_excel_file(file_path, max_items=max_items)
        if ext == ".jsonl":
            with open(file_path, "r", encoding="utf-8") as file:
                for line_number, line in enumerate(file, start=1):
                    line = line.strip()
                    if line:
                        try:
                            item = json.loads(line)
                        except json.JSONDecodeError as exc:
                            invalid_record_count += 1
                            if len(read_errors) < 20:
                                read_errors.append(
                                    f"JSONL 第 {line_number} 行无效: {exc.msg}"
                                )
                            continue
                        total += 1
                        if len(items) < max_items:
                            items.append(item)
        elif ext == ".json":
            file_size = os.path.getsize(file_path)
            if file_size > MAX_JSON_FILE_BYTES:
                return {
                    "count": 0,
                    "returned_count": 0,
                    "truncated": True,
                    "items": [],
                    "error": (
                        "JSON 文件过大，拒绝整体载入内存；"
                        "请改用 JSONL、CSV 或 Excel"
                    ),
                }
            with open(file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
                source_items = data if isinstance(data, list) else [data]
                total = len(source_items)
                items = source_items[:max_items]
        elif ext == ".csv":
            with open(file_path, "r", encoding="utf-8-sig") as file:
                for item in csv.DictReader(file):
                    total += 1
                    if len(items) < max_items:
                        items.append(item)
        else:
            return {
                "count": 0,
                "returned_count": 0,
                "truncated": False,
                "items": [],
                "error": f"不支持的文件类型: {ext}",
            }
    except Exception as exc:
        return {
            "count": 0,
            "returned_count": 0,
            "truncated": False,
            "items": [],
            "error": f"读取失败: {type(exc).__name__}: {exc}",
        }

    result = {
        "count": total,
        "returned_count": len(items),
        "truncated": total > len(items),
        "items": items,
    }
    if read_errors:
        result["error"] = (
            f"发现 {invalid_record_count} 条无效 JSONL 记录: "
            + "; ".join(read_errors)
        )
    return result


def _display_path(file_path: str) -> str:
    try:
        return os.path.relpath(file_path, PROJECT_ROOT)
    except ValueError:
        return str(Path(file_path).resolve())


def _merge_data_files(paths: List[str], max_items: int) -> Dict[str, Any]:
    total = 0
    items: List[Any] = []
    errors: List[str] = []
    sheets: Dict[str, Dict[str, Any]] = {}

    for path in paths:
        data = read_data_file(path, max_items=max_items)
        total += int(data.get("count", 0))
        remaining = max(0, max_items - len(items))
        items.extend(data.get("items", [])[:remaining])
        if data.get("error"):
            errors.append(f"{_display_path(path)}: {data['error']}")
        for sheet_name, sheet_data in data.get("sheets", {}).items():
            merged_sheet = sheets.setdefault(
                sheet_name,
                {"count": 0, "items": []},
            )
            merged_sheet["count"] += int(sheet_data.get("count", 0))
            sheet_remaining = max(0, max_items - len(merged_sheet["items"]))
            merged_sheet["items"].extend(
                sheet_data.get("items", [])[:sheet_remaining]
            )

    for sheet_data in sheets.values():
        sheet_data["returned_count"] = len(sheet_data["items"])
        sheet_data["truncated"] = (
            sheet_data["count"] > sheet_data["returned_count"]
        )

    result: Dict[str, Any] = {
        "file_path": _display_path(paths[-1]),
        "file_paths": [_display_path(path) for path in paths],
        "count": total,
        "returned_count": len(items),
        "truncated": total > len(items),
        "items": items,
    }
    if sheets:
        result["sheets"] = sheets
    if errors:
        result["errors"] = errors
    return result


def get_data_summary(
    platform: str,
    crawler_type: str,
    file_type: str = "jsonl",
    data_root: Optional[str] = None,
) -> Dict[str, Any]:
    """获取指定范围内爬取数据的文件、计数和前三条预览。"""
    groups = _find_data_file_groups(
        platform,
        crawler_type,
        file_type,
        data_root=data_root,
    )
    result: Dict[str, Any] = {
        "platform": platform,
        "crawler_type": crawler_type,
        "files": {},
        "total_count": 0,
    }

    for item_type, paths in groups.items():
        data = _merge_data_files(paths, max_items=3)
        result["total_count"] += data["count"]
        entry = {
            "file_path": data["file_path"],
            "file_paths": data["file_paths"],
            "count": data["count"],
            "preview": data["items"],
        }
        if "sheets" in data:
            entry["sheets"] = data["sheets"]
        if "errors" in data:
            entry["errors"] = data["errors"]
        result["files"][item_type] = entry

    return result


def get_full_data(
    platform: str,
    crawler_type: str,
    file_type: str = "jsonl",
    max_items: int = 200,
    data_root: Optional[str] = None,
) -> Dict[str, Any]:
    """获取指定范围内的数据，每类最多返回 ``max_items`` 条。"""
    groups = _find_data_file_groups(
        platform,
        crawler_type,
        file_type,
        data_root=data_root,
    )
    result: Dict[str, Any] = {
        "platform": platform,
        "crawler_type": crawler_type,
        "files": {},
        "total_count": 0,
    }

    for item_type, paths in groups.items():
        data = _merge_data_files(paths, max_items=max_items)
        result["total_count"] += data["count"]
        result["files"][item_type] = data

    return result
